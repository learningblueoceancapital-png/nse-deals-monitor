#!/usr/bin/env python3
"""
Dividend Initiation Monitor
────────────────────────────
Institutional tracker for rare "dividend initiation" events on NSE/BSE:
companies paying a dividend for the first time in 5 / 10 / 15 years, or
ever since listing. Ignores regular dividend payers.

Data flow
  1. DISCOVER  — pull all BSE corporate-action rows with a "Dividend" purpose
     in the scan window (whole market, one call per date-chunk).
  2. VERIFY    — for every company that appears, pull its FULL corporate
     action history (BSE, back to 1996) and look for an earlier dividend.
     No earlier dividend + long observable history  -> "First Ever".
     No earlier dividend + short observable history -> "First Ever" (young
     listing). No usable history at all              -> Verification Pending.
     Earlier dividend found -> gap in years decides the 5/10/15-year bucket;
     gap < 5 years -> regular payer, excluded.
  3. ENRICH    — CMP + EPS (payout %) from BSE; Market Cap from Screener.in;
     best-effort NSE corporate-actions cross-check (non-blocking).
  4. STORE     — SQLite master table, insert-only, one row per company.
     Never updated or deleted after insertion (per spec).
  5. EXPORT    — clean sorted Excel workbook + console/CI summary.

Phases (per spec)
  Phase 1 — runs once. Backfills FY2025 (1 Apr 2024) through the first
            execution date. Triggered automatically when dividend_data.db
            has no meta.phase1_completed_at.
  Phase 2 — every subsequent run. Scans only since the last run (falls back
            to 7 days if unknown). Appends newly-qualifying companies only.

Usage:
  python dividend_tracker.py                 # auto: phase 1 once, then phase 2
  python dividend_tracker.py --force-phase1   # re-run full historical backfill
  python dividend_tracker.py --from 2024-04-01 --to 2024-12-31   # manual window
"""

import argparse
import io
import logging
import os
import re
import sqlite3
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
from pathlib import Path

import requests

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("dividend_tracker")

# ── Paths ────────────────────────────────────────────────────────────────────
ROOT     = Path(__file__).parent
DB_PATH  = ROOT / "dividend_data.db"
XLSX_OUT = ROOT / "Dividend_Initiation_Tracker.xlsx"
CSV_OUT  = ROOT / "Dividend_Initiation_Tracker.csv"
RECENT_CSV_OUT = ROOT / "Recent_Listing_First_Dividends.csv"

# ── BSE endpoints (public JSON/CSV API behind api.bseindia.com) ────────────────
BSE_CORPACT_URL   = "https://api.bseindia.com/BseIndiaAPI/api/CorpactCSVDownload/w"
BSE_HEADER_URL    = "https://api.bseindia.com/BseIndiaAPI/api/getScripHeaderData/w"
BSE_COMHEADER_URL = "https://api.bseindia.com/BseIndiaAPI/api/ComHeadernew/w"
SCREENER_URL      = "https://www.screener.in/company/{slug}/consolidated/"
SCREENER_URL_ALT  = "https://www.screener.in/company/{slug}/"

BSE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept":  "application/json, text/plain, */*",
    "Origin":  "https://www.bseindia.com",
    "Referer": "https://www.bseindia.com/corporates/corporate_act.aspx",
}
SCREENER_HEADERS = {"User-Agent": BSE_HEADERS["User-Agent"]}

HIST_START      = "1996-01-01"   # floor for BSE electronic corporate-action records
TIMEOUT          = 20
MAX_WORKERS       = 8
REQUEST_SLEEP      = 0.12

FY2025_START = "2024-04-01"

GAP_BUCKETS = [
    (15, "First Dividend in 15 Years"),
    (10, "First Dividend in 10 Years"),
    (5,  "First Dividend in 5 Years"),
]

# Companies listed more recently than this paying their natural first
# dividend are reported separately from the main "rare event" tracker.
RECENT_LISTING_YEARS = 5


# ══════════════════════════════════════════════════════════════════════════════
# Database
# ══════════════════════════════════════════════════════════════════════════════

def open_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS dividend_initiations (
            bse_code           TEXT PRIMARY KEY,
            company_name       TEXT NOT NULL,
            nse_symbol_guess   TEXT,
            dividend_rs        REAL,
            dividend_type      TEXT,
            payout_pct         REAL,
            classification     TEXT NOT NULL,
            cmp                REAL,
            market_cap         TEXT,
            announcement_date  TEXT NOT NULL,
            gap_years          REAL,
            last_dividend_before TEXT,
            citation           TEXT,
            sources_checked    TEXT,
            added_run_date     TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS recent_listing_dividends (
            bse_code           TEXT PRIMARY KEY,
            company_name       TEXT NOT NULL,
            dividend_rs        REAL,
            dividend_type      TEXT,
            payout_pct         REAL,
            cmp                REAL,
            market_cap         TEXT,
            announcement_date  TEXT NOT NULL,
            years_since_listing REAL,
            citation           TEXT,
            added_run_date     TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS verification_pending (
            bse_code           TEXT,
            company_name       TEXT,
            dividend_rs        REAL,
            announcement_date  TEXT,
            reason             TEXT,
            added_run_date     TEXT,
            PRIMARY KEY (bse_code, announcement_date)
        )
    """)
    conn.execute("CREATE TABLE IF NOT EXISTS meta (key TEXT PRIMARY KEY, value TEXT)")
    conn.commit()
    return conn


def get_meta(conn, key, default=None):
    row = conn.execute("SELECT value FROM meta WHERE key=?", (key,)).fetchone()
    return row["value"] if row else default


def set_meta(conn, key, value):
    conn.execute(
        "INSERT INTO meta (key, value) VALUES (?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, value),
    )
    conn.commit()


def already_recorded(conn, bse_code) -> bool:
    return conn.execute(
        "SELECT 1 FROM dividend_initiations WHERE bse_code=? "
        "UNION SELECT 1 FROM recent_listing_dividends WHERE bse_code=?",
        (bse_code, bse_code),
    ).fetchone() is not None


# ══════════════════════════════════════════════════════════════════════════════
# BSE HTTP helpers
# ══════════════════════════════════════════════════════════════════════════════

_session = None


def get_session() -> requests.Session:
    global _session
    if _session is None:
        _session = requests.Session()
        _session.headers.update(BSE_HEADERS)
    return _session


def _get(url: str, retries: int = 3) -> requests.Response | None:
    s = get_session()
    for attempt in range(1, retries + 1):
        try:
            r = s.get(url, timeout=TIMEOUT)
            if r.status_code == 200:
                return r
            log.debug("Non-200 (%s) on %s", r.status_code, url)
        except Exception as exc:
            log.debug("Attempt %d/%d failed for %s — %s", attempt, retries, url, exc)
        time.sleep(attempt * 0.6)
    return None


def _bse_date(d: str) -> str:
    """YYYY-MM-DD -> YYYYMMDD (BSE API format)."""
    return d.replace("-", "")


def fetch_corpact_window(fdate: str, tdate: str, scripcode: str = "") -> list[dict]:
    """Pull corporate-action rows for the whole market (scripcode='') or one company."""
    url = (
        f"{BSE_CORPACT_URL}?scripcode={scripcode}"
        f"&Fdate={_bse_date(fdate)}&Tdate={_bse_date(tdate)}"
        f"&Purposecode=&strSearch=S"
    )
    resp = _get(url)
    time.sleep(REQUEST_SLEEP)
    if resp is None:
        return []
    lines = resp.text.splitlines()
    if len(lines) < 2:
        return []
    rows = []
    for line in lines[1:]:
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 5:
            continue
        rows.append({
            "bse_code":     parts[0],
            "security_name": parts[1],
            "company_name": parts[2],
            "ex_date":      parts[3],
            "purpose":      parts[4],
        })
    return rows


def discover_window(fdate: str, tdate: str) -> list[dict]:
    """Chunk the discovery scan into <=6-month windows (defensive against any
    server-side truncation) and merge, de-duplicated by (bse_code, ex_date, purpose)."""
    chunks = []
    start = datetime.strptime(fdate, "%Y-%m-%d")
    end   = datetime.strptime(tdate, "%Y-%m-%d")
    cur = start
    while cur <= end:
        chunk_end = min(cur + timedelta(days=182), end)
        chunks.append((cur.strftime("%Y-%m-%d"), chunk_end.strftime("%Y-%m-%d")))
        cur = chunk_end + timedelta(days=1)

    seen = set()
    all_rows = []
    for cf, ct in chunks:
        log.info("Discovery scan %s -> %s", cf, ct)
        for row in fetch_corpact_window(cf, ct):
            if not is_dividend_purpose(row["purpose"]):
                continue
            key = (row["bse_code"], row["ex_date"], row["purpose"])
            if key in seen:
                continue
            seen.add(key)
            all_rows.append(row)
    return all_rows


DIVIDEND_RE = re.compile(
    r"(Interim Dividend|Final Dividend|Special Dividend|Dividend)\s*-\s*Rs\.?\s*-?\s*([\d.]+)",
    re.IGNORECASE,
)


def is_dividend_purpose(purpose: str) -> bool:
    p = purpose.lower()
    return "dividend" in p and "no dividend" not in p


def parse_dividend_amount(purpose: str) -> tuple[str, float | None]:
    m = DIVIDEND_RE.search(purpose)
    if m:
        dtype = m.group(1).title()
        try:
            return dtype, round(float(m.group(2)), 4)
        except ValueError:
            return dtype, None
    dtype = "Special Dividend" if "special" in purpose.lower() else "Dividend"
    return dtype, None


def parse_bse_date(raw: str) -> str | None:
    raw = raw.strip()
    for fmt in ("%d %b %Y", "%d-%b-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


# ══════════════════════════════════════════════════════════════════════════════
# Per-company history check & classification
# ══════════════════════════════════════════════════════════════════════════════
#
# IMPORTANT — data-quality finding from validation runs:
# BSE's CorpactCSVDownload endpoint, when queried per-scrip over a wide
# (multi-year) date range, is NOT reliably complete. Spot checks turned up
# household-name, decades-long regular dividend payers (Maruti Suzuki,
# Mahindra & Mahindra) for which this endpoint returns almost no historical
# rows — which would have produced false "First Ever Dividend" positives.
# Yahoo Finance's chart API (events=div) proved far more complete/reliable
# in the same spot checks (e.g. it correctly returned Maruti's full 22-year
# annual dividend record). It is therefore used as the PRIMARY source for
# the historical no-prior-dividend check; BSE's per-scrip history is kept
# only as a secondary cross-check that can *strengthen* a finding (an extra
# dividend it catches that Yahoo missed) but never used alone to certify
# a "first ever" / large-gap claim.

_yahoo_candidates_cache: dict[str, list[str]] = {}


def _candidate_symbols(bse_code: str, security_name: str, company_name: str) -> list[str]:
    """Symbol variants to try, .NS first — spot checks showed Yahoo's NSE
    (.NS) feed is consistently the most complete for dividend history;
    numeric-scripcode .BO in particular can silently return near-empty
    dividend data even for major, decades-long regular payers."""
    if bse_code in _yahoo_candidates_cache:
        return _yahoo_candidates_cache[bse_code]

    direct = [f"{security_name}.NS", f"{security_name}.BO", f"{bse_code}.BO"]
    clean_name = re.sub(r"\s*(Ltd\.?|Limited)\s*$", "", company_name, flags=re.IGNORECASE).strip()
    searched = []
    for query in (clean_name, security_name):
        try:
            r = requests.get(
                "https://query2.finance.yahoo.com/v1/finance/search",
                params={"q": query, "quotesCount": 5, "newsCount": 0},
                headers=SCREENER_HEADERS, timeout=10,
            )
            quotes = r.json().get("quotes", []) if r.status_code == 200 else []
            searched += [q["symbol"] for q in quotes if q.get("exchange") in ("NSI", "BSE")]
        except Exception:
            pass
        time.sleep(0.05)
        if searched:
            break

    seen, ordered = set(), []
    for sym in direct + searched:
        if sym not in seen:
            seen.add(sym)
            ordered.append(sym)
    _yahoo_candidates_cache[bse_code] = ordered
    return ordered


def _fetch_yahoo_events(symbol: str) -> dict | None:
    try:
        r = requests.get(
            f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}",
            params={"period1": 0, "period2": 9_999_999_999, "interval": "1mo", "events": "div"},
            headers=SCREENER_HEADERS, timeout=15,
        )
        if r.status_code != 200:
            return None
        return r.json()["chart"]["result"][0]
    except Exception:
        return None


def fetch_yahoo_dividend_history(bse_code: str, security_name: str, company_name: str,
                                  before_date: str) -> dict | None:
    """
    Merges dividend-events data across every Yahoo symbol variant that
    resolves for this company (not just the first hit) — data completeness
    varies by symbol variant even for the same company, so the union of
    what every reachable variant shows is the most trustworthy signal.
    Returns None only if NO variant resolved at all.
    Otherwise:
      last_dividend_date : most recent dividend Ex-Date strictly before
                            before_date across all variants, or None
      first_trade_date   : earliest known listing date across variants
    """
    cutoff = datetime.strptime(before_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    any_resolved = False
    all_div_dates, all_first_trade = [], []

    for sym in _candidate_symbols(bse_code, security_name, company_name):
        result = _fetch_yahoo_events(sym)
        time.sleep(0.05)
        if result is None:
            continue
        any_resolved = True

        first_trade_ts = result.get("meta", {}).get("firstTradeDate")
        if first_trade_ts:
            all_first_trade.append(datetime.fromtimestamp(first_trade_ts, tz=timezone.utc))

        for ev in result.get("events", {}).get("dividends", {}).values():
            d = datetime.fromtimestamp(ev["date"], tz=timezone.utc)
            if d < cutoff:
                all_div_dates.append(d)

    if not any_resolved:
        return None

    return {
        "last_dividend_date": max(all_div_dates).strftime("%Y-%m-%d") if all_div_dates else None,
        "first_trade_date": min(all_first_trade).strftime("%Y-%m-%d") if all_first_trade else None,
    }


def check_history_bse(bse_code: str, before_date: str) -> str | None:
    """Secondary cross-check only — returns the most recent prior dividend
    Ex-Date this BSE endpoint shows (may be incomplete; never treat its
    silence as proof of absence)."""
    day_before = (datetime.strptime(before_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    rows = fetch_corpact_window(HIST_START, day_before, scripcode=bse_code)
    dates = [parse_bse_date(r["ex_date"]) for r in rows if is_dividend_purpose(r["purpose"])]
    dates = [d for d in dates if d]
    return max(dates) if dates else None


def check_history(bse_code: str, security_name: str, company_name: str, before_date: str) -> dict:
    yahoo = fetch_yahoo_dividend_history(bse_code, security_name, company_name, before_date)
    bse_last_div = check_history_bse(bse_code, before_date)

    candidates = [d for d in [
        yahoo["last_dividend_date"] if yahoo else None,
        bse_last_div,
    ] if d]

    return {
        "last_dividend_date": max(candidates) if candidates else None,
        "first_trade_date": yahoo["first_trade_date"] if yahoo else None,
        "yahoo_verified": yahoo is not None,
    }


def classify(announcement_date: str, hist: dict) -> tuple[str | None, float | None, str | None, str | None]:
    """
    Returns (classification, gap_years, last_dividend_before, pending_reason).
    classification is None (and pending_reason set) when the company should
    be excluded pending manual verification. classification is also None
    (pending_reason=None) when the company is a regular payer (gap < 5y) —
    caller should simply skip these, no need to log them.
    """
    ann = datetime.strptime(announcement_date, "%Y-%m-%d")
    last_div = hist["last_dividend_date"]
    first_trade_date = hist["first_trade_date"]

    if last_div:
        gap_years = round((ann - datetime.strptime(last_div, "%Y-%m-%d")).days / 365.25, 1)
        for threshold, label in GAP_BUCKETS:
            if gap_years >= threshold:
                return label, gap_years, last_div, None
        return None, gap_years, last_div, None  # regular payer — skip silently

    # No prior dividend found. Only trust this "first ever" claim when we
    # got a clean Yahoo Finance read (a materially complete, third-party
    # dividend-events feed) with a known listing date — never on BSE's
    # per-scrip history alone (proven incomplete in spot checks).
    if hist["yahoo_verified"] and first_trade_date:
        years_listed = round((ann - datetime.strptime(first_trade_date, "%Y-%m-%d")).days / 365.25, 1)
        if years_listed < RECENT_LISTING_YEARS:
            # A company that only listed recently paying its natural first
            # dividend isn't a "rare" event the same way an established
            # company's first-ever dividend is — kept out of the main
            # tracker, reported separately instead.
            return "RECENT_LISTING", years_listed, None, None
        return "First Ever Dividend Since Listing", years_listed, None, None

    return None, None, None, (
        "Could not obtain a verified complete dividend history for this company "
        "(market-data symbol unresolved or listing date unavailable) — "
        "verify manually (Screener/Moneycontrol/annual report) before treating as a first-ever event"
    )


# ══════════════════════════════════════════════════════════════════════════════
# Enrichment: CMP, EPS/payout, market cap
# ══════════════════════════════════════════════════════════════════════════════

def fetch_cmp(bse_code: str) -> float | None:
    url = f"{BSE_HEADER_URL}?Debtflag=&scripcode={bse_code}&seriesid="
    resp = _get(url, retries=2)
    time.sleep(REQUEST_SLEEP)
    if not resp:
        return None
    try:
        ltp = resp.json().get("CurrRate", {}).get("LTP")
        return round(float(ltp), 2) if ltp else None
    except Exception:
        return None


def fetch_eps(bse_code: str) -> float | None:
    url = f"{BSE_COMHEADER_URL}?quotetype=EQ&scripcode={bse_code}&seriesid="
    resp = _get(url, retries=2)
    time.sleep(REQUEST_SLEEP)
    if not resp:
        return None
    try:
        eps = resp.json().get("EPS")
        return round(float(eps), 2) if eps and eps not in ("", "0") else None
    except Exception:
        return None


def fetch_market_cap(security_name: str) -> str:
    """Best-effort Market Cap (₹ Cr) from Screener.in, keyed off the BSE
    ticker (usually identical/close to the NSE symbol). Returns 'N/A' on
    any failure — never blocks the pipeline."""
    slug = security_name.strip().upper()
    for tmpl in (SCREENER_URL, SCREENER_URL_ALT):
        try:
            r = requests.get(tmpl.format(slug=slug), headers=SCREENER_HEADERS, timeout=15)
            if r.status_code != 200:
                continue
            m = re.search(r'Market Cap.*?<span class="(?:number|value)">([\d,]+)</span>', r.text, re.S)
            if m:
                return f"Rs {m.group(1)} Cr"
        except Exception:
            pass
        time.sleep(REQUEST_SLEEP)
    return "N/A"


# ══════════════════════════════════════════════════════════════════════════════
# NSE cross-check (best-effort, non-blocking secondary source)
# ══════════════════════════════════════════════════════════════════════════════

def nse_cross_check_symbols(fdate: str, tdate: str) -> set[str] | None:
    """
    Returns a set of company names NSE also shows a dividend corporate
    action for in this window, or None if NSE could not be reached at all
    (e.g. Akamai bot-protection blocked this runner). Never raises.
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log.warning("Playwright not installed — skipping NSE cross-check")
        return None

    nse_fmt = lambda d: datetime.strptime(d, "%Y-%m-%d").strftime("%d-%m-%Y")
    url = (
        "https://www.nseindia.com/api/corporates-corporateActions"
        f"?index=equities&from_date={nse_fmt(fdate)}&to_date={nse_fmt(tdate)}"
    )
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-blink-features=AutomationControlled",
                      "--disable-dev-shm-usage"],
            )
            ctx = browser.new_context(
                user_agent=BSE_HEADERS["User-Agent"], locale="en-IN", timezone_id="Asia/Kolkata",
            )
            page = ctx.new_page()
            page.goto("https://www.nseindia.com", wait_until="domcontentloaded", timeout=25_000)
            page.wait_for_timeout(2000)
            resp = page.goto(url, timeout=20_000)
            if resp.status != 200:
                browser.close()
                return None
            data = resp.json()
            browser.close()
            names = {
                str(item.get("comp", "")).strip().upper()
                for item in data
                if is_dividend_purpose(str(item.get("subject", "")))
            }
            return names or None
    except Exception as exc:
        log.warning("NSE cross-check unavailable this run: %s", exc)
        return None


# ══════════════════════════════════════════════════════════════════════════════
# Pipeline
# ══════════════════════════════════════════════════════════════════════════════

def process_company(row: dict, run_date: str) -> tuple[str, dict]:
    """Returns ('qualified'|'recent_listing'|'pending'|'skip', record_dict)."""
    bse_code = row["bse_code"]
    ann_date = parse_bse_date(row["ex_date"])
    if not ann_date:
        return "skip", {}

    dtype, amount = parse_dividend_amount(row["purpose"])
    hist = check_history(bse_code, row["security_name"], row["company_name"], ann_date)
    classification, gap_years, last_before, pending_reason = classify(ann_date, hist)

    if pending_reason:
        return "pending", {
            "bse_code": bse_code, "company_name": row["company_name"],
            "dividend_rs": amount, "announcement_date": ann_date,
            "reason": pending_reason, "added_run_date": run_date,
        }
    if classification is None:
        return "skip", {}  # regular payer

    cmp_ = fetch_cmp(bse_code)
    eps = fetch_eps(bse_code)
    payout_pct = round((amount / eps) * 100, 1) if (amount and eps and eps > 0) else None
    market_cap = fetch_market_cap(row["security_name"])
    citation = (f"BSE Corporate Actions — scrip {bse_code}, ex-date {ann_date} "
                f"(https://www.bseindia.com/corporates/corporate_act.aspx)")

    if classification == "RECENT_LISTING":
        return "recent_listing", {
            "bse_code": bse_code,
            "company_name": row["company_name"],
            "dividend_rs": amount,
            "dividend_type": dtype,
            "payout_pct": payout_pct,
            "cmp": cmp_,
            "market_cap": market_cap,
            "announcement_date": ann_date,
            "years_since_listing": gap_years,
            "citation": citation,
            "added_run_date": run_date,
        }

    return "qualified", {
        "bse_code": bse_code,
        "company_name": row["company_name"],
        "nse_symbol_guess": row["security_name"],
        "dividend_rs": amount,
        "dividend_type": dtype,
        "payout_pct": payout_pct,
        "classification": classification,
        "cmp": cmp_,
        "market_cap": market_cap,
        "announcement_date": ann_date,
        "gap_years": gap_years,
        "last_dividend_before": last_before,
        "citation": citation,
        "added_run_date": run_date,
    }


def run_pipeline(fdate: str, tdate: str, conn: sqlite3.Connection) -> dict:
    run_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    log.info("Scan window: %s -> %s", fdate, tdate)

    raw_rows = discover_window(fdate, tdate)
    log.info("Dividend-purpose rows found: %d", len(raw_rows))

    # One candidate per company: earliest dividend row in the window.
    by_company: dict[str, dict] = {}
    for row in raw_rows:
        d = parse_bse_date(row["ex_date"])
        if not d:
            continue
        existing = by_company.get(row["bse_code"])
        if existing is None or d < parse_bse_date(existing["ex_date"]):
            by_company[row["bse_code"]] = row

    candidates = [r for code, r in by_company.items() if not already_recorded(conn, code)]
    log.info("Unique companies to evaluate (not already recorded): %d", len(candidates))

    nse_names = nse_cross_check_symbols(fdate, tdate)
    sources_note = "BSE + NSE" if nse_names is not None else "BSE (NSE cross-check unavailable this run)"

    qualified, recent_listings, pending, scanned = [], [], [], 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(process_company, row, run_date): row for row in candidates}
        for fut in as_completed(futures):
            scanned += 1
            try:
                kind, record = fut.result()
            except Exception as exc:
                log.warning("Error processing %s: %s", futures[fut].get("company_name"), exc)
                continue
            if kind == "qualified":
                if nse_names is not None:
                    record["sources_checked"] = (
                        "BSE + NSE" if record["company_name"].strip().upper() in nse_names else "BSE only"
                    )
                else:
                    record["sources_checked"] = sources_note
                qualified.append(record)
            elif kind == "recent_listing":
                recent_listings.append(record)
            elif kind == "pending":
                pending.append(record)

    for rec in qualified:
        conn.execute("""
            INSERT OR IGNORE INTO dividend_initiations (
                bse_code, company_name, nse_symbol_guess, dividend_rs, dividend_type,
                payout_pct, classification, cmp, market_cap, announcement_date,
                gap_years, last_dividend_before, citation, sources_checked, added_run_date
            ) VALUES (
                :bse_code, :company_name, :nse_symbol_guess, :dividend_rs, :dividend_type,
                :payout_pct, :classification, :cmp, :market_cap, :announcement_date,
                :gap_years, :last_dividend_before, :citation, :sources_checked, :added_run_date
            )
        """, rec)
    for rec in recent_listings:
        conn.execute("""
            INSERT OR IGNORE INTO recent_listing_dividends (
                bse_code, company_name, dividend_rs, dividend_type, payout_pct,
                cmp, market_cap, announcement_date, years_since_listing, citation, added_run_date
            ) VALUES (
                :bse_code, :company_name, :dividend_rs, :dividend_type, :payout_pct,
                :cmp, :market_cap, :announcement_date, :years_since_listing, :citation, :added_run_date
            )
        """, rec)
    for rec in pending:
        conn.execute("""
            INSERT OR IGNORE INTO verification_pending
                (bse_code, company_name, dividend_rs, announcement_date, reason, added_run_date)
            VALUES (:bse_code, :company_name, :dividend_rs, :announcement_date, :reason, :added_run_date)
        """, rec)
    conn.commit()

    return {
        "companies_scanned": len(candidates),
        "announcements_reviewed": len(raw_rows),
        "new_qualifying": len(qualified),
        "new_recent_listings": len(recent_listings),
        "verification_pending": len(pending),
        "qualified_names": [q["company_name"] for q in qualified],
    }


# ══════════════════════════════════════════════════════════════════════════════
# Export
# ══════════════════════════════════════════════════════════════════════════════

CLASS_ORDER = {
    "First Ever Dividend Since Listing": 0,
    "First Dividend in 15 Years": 1,
    "First Dividend in 10 Years": 2,
    "First Dividend in 5 Years": 3,
}


def sorted_master_rows(conn: sqlite3.Connection) -> list:
    """Master table rows in required sort order: classification bucket
    (First Ever -> 15y -> 10y -> 5y), then announcement date latest-first
    within each bucket."""
    rows = conn.execute("""
        SELECT company_name, dividend_rs, payout_pct, classification, cmp,
               announcement_date, market_cap, citation
        FROM dividend_initiations
    """).fetchall()
    grouped: dict[str, list] = {}
    for r in rows:
        grouped.setdefault(r["classification"], []).append(r)
    final_rows = []
    for cls in sorted(grouped, key=lambda c: CLASS_ORDER.get(c, 9)):
        final_rows.extend(sorted(grouped[cls], key=lambda r: r["announcement_date"], reverse=True))
    return final_rows


HEADERS = ["Company", "Dividend Declared (Rs/Share)", "Dividend Payout %",
           "Classification", "CMP", "Announcement Date", "Market Cap", "Source"]


RECENT_HEADERS = ["Company", "Dividend Declared (Rs/Share)", "Dividend Payout %",
                  "CMP", "Announcement Date", "Market Cap", "Years Since Listing", "Source"]


def export_csv(conn: sqlite3.Connection) -> None:
    """Git-friendly plaintext master table — this is the file committed to
    the repo each run (this repo's .gitignore excludes *.xlsx binaries)."""
    import csv
    final_rows = sorted_master_rows(conn)
    with open(CSV_OUT, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for r in final_rows:
            w.writerow([
                r["company_name"], r["dividend_rs"],
                r["payout_pct"] if r["payout_pct"] is not None else "N/A",
                r["classification"], r["cmp"] if r["cmp"] is not None else "N/A",
                r["announcement_date"], r["market_cap"], r["citation"],
            ])
    log.info("Exported %d qualifying rows -> %s", len(final_rows), CSV_OUT)

    recent_rows = conn.execute("""
        SELECT company_name, dividend_rs, payout_pct, cmp, market_cap,
               announcement_date, years_since_listing, citation
        FROM recent_listing_dividends ORDER BY announcement_date DESC
    """).fetchall()
    with open(RECENT_CSV_OUT, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(RECENT_HEADERS)
        for r in recent_rows:
            w.writerow([
                r["company_name"], r["dividend_rs"],
                r["payout_pct"] if r["payout_pct"] is not None else "N/A",
                r["cmp"] if r["cmp"] is not None else "N/A", r["announcement_date"],
                r["market_cap"], r["years_since_listing"], r["citation"],
            ])
    log.info("Exported %d recent-listing first-dividend rows -> %s", len(recent_rows), RECENT_CSV_OUT)


CLASS_FILL = {
    "First Ever Dividend Since Listing": "BDD7EE",
    "First Dividend in 15 Years":        "D9E8F5",
    "First Dividend in 10 Years":        "E8F1FA",
    "First Dividend in 5 Years":         "F2F8FC",
}

BSE_CORPACT_PAGE = "https://www.bseindia.com/corporates/corporate_act.aspx"


def _style_workbook_common():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return {
        "hdr_fill":  PatternFill("solid", fgColor="1F4E79"),
        "hdr_font":  Font(bold=True, color="FFFFFF", name="Calibri", size=11),
        "title_font": Font(bold=True, color="1F4E79", name="Calibri", size=14),
        "sub_font":  Font(italic=True, color="5A6B7B", name="Calibri", size=9),
        "body_font": Font(name="Calibri", size=10),
        "link_font": Font(name="Calibri", size=10, color="1155CC", underline="single"),
        "border":    Border(*[Side(style="thin", color="C7D3DE")] * 4),
        "center":    Alignment(horizontal="center", vertical="center"),
        "left":      Alignment(horizontal="left", vertical="center"),
        "right":     Alignment(horizontal="right", vertical="center"),
    }


def _write_title_block(ws, sty, title: str, subtitle: str, ncols: int):
    from openpyxl.utils import get_column_letter
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=f"BlueOcean Capital — {title}")
    c.font = sty["title_font"]
    c.alignment = sty["left"]
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = sty["sub_font"]
    c2.alignment = sty["left"]
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[3].height = 6  # spacer


def _write_header_row(ws, sty, headers, row: int):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = sty["hdr_fill"]
        cell.font = sty["hdr_font"]
        cell.alignment = sty["center"]
        cell.border = sty["border"]
    ws.row_dimensions[row].height = 20


def export_xlsx(conn: sqlite3.Connection) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    generated = datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC")
    sty = _style_workbook_common()
    wb = Workbook()

    # ── Sheet 1: Main tracker ──────────────────────────────────────────
    final_rows = sorted_master_rows(conn)
    ws = wb.active
    ws.title = "Dividend Initiation Tracker"
    HDR_ROW = 4
    ncols = len(HEADERS)
    _write_title_block(ws, sty, "Dividend Initiation Monitor",
                        f"Companies paying a dividend after a 5/10/15-year gap, or for the first "
                        f"time since listing  |  {len(final_rows)} companies  |  generated {generated}",
                        ncols)
    _write_header_row(ws, sty, HEADERS, HDR_ROW)

    NUM_FMT_RS = '#,##0.00'
    NUM_FMT_PCT = '0.0"%"'

    for i, r in enumerate(final_rows, start=HDR_ROW + 1):
        cls = r["classification"]
        row_fill_hex = CLASS_FILL.get(cls, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=row_fill_hex)

        cells = [
            (r["company_name"], sty["left"], None),
            (r["dividend_rs"], sty["right"], NUM_FMT_RS if r["dividend_rs"] is not None else None),
            (r["payout_pct"] if r["payout_pct"] is not None else "N/A", sty["right"],
             NUM_FMT_PCT if r["payout_pct"] is not None else None),
            (cls, sty["left"], None),
            (r["cmp"] if r["cmp"] is not None else "N/A", sty["right"],
             NUM_FMT_RS if r["cmp"] is not None else None),
            (r["announcement_date"], sty["center"], None),
            (r["market_cap"], sty["right"], None),
            ("BSE ↗", sty["center"], None),
        ]
        for c, (val, align, numfmt) in enumerate(cells, start=1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.border = sty["border"]
            cell.alignment = align
            cell.fill = row_fill
            if c == ncols:
                cell.font = sty["link_font"]
                cell.hyperlink = BSE_CORPACT_PAGE
            else:
                cell.font = sty["body_font"]
            if numfmt:
                cell.number_format = numfmt

    widths = [32, 15, 13, 26, 11, 15, 14, 9]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = f"A{HDR_ROW + 1}"
    ws.auto_filter.ref = f"A{HDR_ROW}:{get_column_letter(ncols)}{max(HDR_ROW, len(final_rows) + HDR_ROW)}"
    ws.sheet_view.showGridLines = False

    # ── Sheet 2: Recent Listings ────────────────────────────────────────
    recent_rows = conn.execute("""
        SELECT company_name, dividend_rs, payout_pct, cmp, market_cap,
               announcement_date, years_since_listing, citation
        FROM recent_listing_dividends ORDER BY announcement_date DESC
    """).fetchall()
    ws3 = wb.create_sheet("Recent Listings")
    n3 = len(RECENT_HEADERS)
    _write_title_block(ws3, sty, "Recent Listings — First Dividend",
                        f"Companies whose first-ever dividend came within {RECENT_LISTING_YEARS} years "
                        f"of listing — expected, not a rare event; kept separate from the main tracker  |  "
                        f"{len(recent_rows)} companies", n3)
    _write_header_row(ws3, sty, RECENT_HEADERS, HDR_ROW)
    for i, r in enumerate(recent_rows, start=HDR_ROW + 1):
        cells = [
            (r["company_name"], sty["left"], None),
            (r["dividend_rs"], sty["right"], NUM_FMT_RS if r["dividend_rs"] is not None else None),
            (r["payout_pct"] if r["payout_pct"] is not None else "N/A", sty["right"],
             NUM_FMT_PCT if r["payout_pct"] is not None else None),
            (r["cmp"] if r["cmp"] is not None else "N/A", sty["right"],
             NUM_FMT_RS if r["cmp"] is not None else None),
            (r["announcement_date"], sty["center"], None),
            (r["market_cap"], sty["right"], None),
            (r["years_since_listing"], sty["right"], '0.0'),
            ("BSE ↗", sty["center"], None),
        ]
        for c, (val, align, numfmt) in enumerate(cells, start=1):
            cell = ws3.cell(row=i, column=c, value=val)
            cell.border = sty["border"]
            cell.alignment = align
            if c == n3:
                cell.font = sty["link_font"]
                cell.hyperlink = BSE_CORPACT_PAGE
            else:
                cell.font = sty["body_font"]
            if numfmt:
                cell.number_format = numfmt
    for c, w in enumerate([32, 15, 13, 11, 15, 14, 15, 9], start=1):
        ws3.column_dimensions[get_column_letter(c)].width = w
    ws3.freeze_panes = f"A{HDR_ROW + 1}"
    ws3.auto_filter.ref = f"A{HDR_ROW}:{get_column_letter(n3)}{max(HDR_ROW, len(recent_rows) + HDR_ROW)}"
    ws3.sheet_view.showGridLines = False

    # ── Sheet 3: Verification Pending ───────────────────────────────────
    pending_rows = conn.execute("""
        SELECT company_name, dividend_rs, announcement_date, reason, added_run_date
        FROM verification_pending ORDER BY added_run_date DESC
    """).fetchall()
    ws2 = wb.create_sheet("Verification Pending")
    pending_headers = ["Company", "Dividend Declared (Rs/Share)", "Announcement Date", "Reason", "Flagged On"]
    n2 = len(pending_headers)
    _write_title_block(ws2, sty, "Verification Pending",
                        f"Dividend history could not be independently verified — excluded from the "
                        f"tracker until manually confirmed  |  {len(pending_rows)} companies", n2)
    _write_header_row(ws2, sty, pending_headers, HDR_ROW)
    for i, r in enumerate(pending_rows, start=HDR_ROW + 1):
        vals = [r["company_name"], r["dividend_rs"], r["announcement_date"], r["reason"], r["added_run_date"]]
        aligns = [sty["left"], sty["right"], sty["center"], sty["left"], sty["center"]]
        for c, (val, align) in enumerate(zip(vals, aligns), start=1):
            cell = ws2.cell(row=i, column=c, value=val)
            cell.border = sty["border"]
            cell.font = sty["body_font"]
            cell.alignment = align
            if c == 2 and val is not None:
                cell.number_format = NUM_FMT_RS
    for c, w in enumerate([30, 16, 16, 55, 14], start=1):
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.freeze_panes = f"A{HDR_ROW + 1}"
    ws2.sheet_view.showGridLines = False

    wb.save(XLSX_OUT)
    log.info("Exported %d qualifying + %d recent-listing + %d pending rows -> %s",
              len(final_rows), len(recent_rows), len(pending_rows), XLSX_OUT)


# ══════════════════════════════════════════════════════════════════════════════
# Summary
# ══════════════════════════════════════════════════════════════════════════════

def print_summary(stats: dict) -> None:
    lines = ["## Weekly Summary", ""]
    lines.append(f"- Companies Scanned: {stats['companies_scanned']}")
    lines.append(f"- Dividend Announcements Reviewed: {stats['announcements_reviewed']}")
    lines.append(f"- New Qualifying Companies: {stats['new_qualifying']}")
    lines.append(f"- New Recent-Listing First Dividends (separate note, not in main tracker): "
                 f"{stats['new_recent_listings']}")
    lines.append(f"- Verification Pending: {stats['verification_pending']}")
    lines.append("")
    if stats["new_qualifying"] == 0:
        lines.append("No new qualifying dividend initiations this week.")
    else:
        lines.append("New qualifying companies this run:")
        for name in stats["qualified_names"]:
            lines.append(f"  - {name}")
    text = "\n".join(lines)
    print(text)
    summary_path = os.environ.get("GITHUB_STEP_SUMMARY")
    if summary_path:
        with open(summary_path, "a", encoding="utf-8") as f:
            f.write(text + "\n")


# ══════════════════════════════════════════════════════════════════════════════
# Entry point
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description="Dividend Initiation Monitor")
    ap.add_argument("--force-phase1", action="store_true", help="Re-run full historical backfill")
    ap.add_argument("--from", dest="from_date", help="Manual window start YYYY-MM-DD")
    ap.add_argument("--to", dest="to_date", help="Manual window end YYYY-MM-DD")
    args = ap.parse_args()

    conn = open_db()
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    if args.from_date and args.to_date:
        fdate, tdate = args.from_date, args.to_date
    elif args.force_phase1 or not get_meta(conn, "phase1_completed_at"):
        fdate, tdate = FY2025_START, today
        log.info("PHASE 1 — one-time historical backfill (%s -> %s)", fdate, tdate)
    else:
        fdate = get_meta(conn, "last_run_date")
        if fdate:
            fdate = (datetime.strptime(fdate, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
        else:
            fdate = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d")
        tdate = today
        log.info("PHASE 2 — incremental scan since last run (%s -> %s)", fdate, tdate)

    stats = run_pipeline(fdate, tdate, conn)
    export_csv(conn)
    export_xlsx(conn)

    set_meta(conn, "last_run_date", today)
    if not get_meta(conn, "phase1_completed_at"):
        set_meta(conn, "phase1_completed_at", today)

    print_summary(stats)
    conn.close()


if __name__ == "__main__":
    main()
