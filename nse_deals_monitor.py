#!/usr/bin/env python3
"""
NSE Bulk & Block Deal Monitor — Playwright Edition
───────────────────────────────────────────────────
Watches https://www.nseindia.com/all-reports for bulk and block deal
files published by the exchange. As soon as new data appears for the
current trading session, it is downloaded, processed into a formatted
Excel report, and emailed automatically.

How it works:
  1. Playwright launches a real Chromium browser (bypasses Akamai / Cloudflare).
  2. The browser navigates NSE, picking up all required session cookies.
  3. Network responses are intercepted — when the exchange publishes today's
     deals the API returns non-empty data; that triggers the pipeline.
  4. Cookies are exported to a requests.Session so API polling is cheap
     (no full page reload needed every cycle).
  5. The browser is kept warm and cookies refreshed every COOKIE_TTL seconds.

Env vars (.env or shell):
  GMAIL_USER          your Gmail address
  GMAIL_APP_PASSWORD  16-char Gmail App Password
  EMAIL_TO            comma-separated recipients
  POLL_INTERVAL       seconds between API polls (default 120)
  COOKIE_TTL          seconds before browser refreshes cookies (default 1800)
  OUTPUT_DIR          folder for saved Excel files (default: script dir)
"""

import os
import io
import sys
import time
import signal
import hashlib
import logging
import datetime
import smtplib
import ssl
from pathlib import Path
from email.message import EmailMessage

import pandas as pd
import requests
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("nse_monitor")

# ── Config ─────────────────────────────────────────────────────────────────────
GMAIL_USER     = os.getenv("GMAIL_USER", "")
GMAIL_PASSWORD = os.getenv("GMAIL_APP_PASSWORD", "")
EMAIL_TO       = [e.strip() for e in os.getenv("EMAIL_TO", "research@blueoceancapital.co.in").split(",") if e.strip()]
POLL_INTERVAL  = int(os.getenv("POLL_INTERVAL", "120"))     # seconds between API polls
COOKIE_TTL     = int(os.getenv("COOKIE_TTL",   "1800"))     # seconds before cookie refresh
OUTPUT_DIR     = Path(os.getenv("OUTPUT_DIR", Path(__file__).parent))
IST            = datetime.timezone(datetime.timedelta(hours=5, minutes=30))

# ── NSE URLs ───────────────────────────────────────────────────────────────────
NSE_HOME       = "https://www.nseindia.com"
NSE_ALL_REPORTS= "https://www.nseindia.com/all-reports"
NSE_DEALS_PAGE = "https://www.nseindia.com/report-detail/display-bulk-and-block-deals"
NSE_DEALS_API  = "https://www.nseindia.com/api/historicalOR/bulk-block-short-deals"

# ── Column mapping — CSV download endpoint (returns ALL records, no server cap)
# The JSON API is hard-capped at 70 records; csv=true bypasses that limit.
CSV_COL_MAP = {
    "Date":                           "Date",
    "Security Name":                  "Security Name",
    "Client Name":                    "Client Name",
    "Buy / Sell":                     "Buy/Sell",
    "Quantity Traded":                "Quantity Traded",
    "Trade Price / Wght. Avg. Price": "Traded Price",
}
REQUIRED_COLS = ["Date", "Security Name", "Client Name", "Buy/Sell",
                 "Quantity Traded", "Traded Price"]
OUTPUT_COLS   = REQUIRED_COLS + ["Traded Value"]

# ── Excel style constants ──────────────────────────────────────────────────────
HDR_FILL    = PatternFill("solid", fgColor="1F4E79")
HDR_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
BODY_FONT   = Font(name="Calibri", size=11)
_side       = Side(style="thin", color="A0B4CC")
CELL_BORDER = Border(left=_side, right=_side, top=_side, bottom=_side)
INR_FMT     = r'₹#,##0.00'


# ══════════════════════════════════════════════════════════════════════════════
# Browser session
# ══════════════════════════════════════════════════════════════════════════════

class NSEBrowserSession:
    """
    Manages a persistent Playwright Chromium session.
    Exports cookies to a requests.Session for lightweight API polling.
    Auto-refreshes cookies when COOKIE_TTL expires.
    """

    _BROWSER_ARGS = [
        "--no-sandbox",
        "--disable-blink-features=AutomationControlled",
        "--disable-dev-shm-usage",
        "--disable-infobars",
        "--window-size=1920,1080",
    ]

    def __init__(self):
        self._pw         = None
        self._browser    = None
        self._context    = None
        self._api_session: requests.Session | None = None
        self._cookie_ts  = 0.0   # epoch time of last cookie refresh

    # ── Lifecycle ──────────────────────────────────────────────────────────────

    def start(self) -> None:
        """Launch browser and perform initial cookie seeding."""
        log.info("Launching Chromium …")
        self._pw      = sync_playwright().start()
        self._browser = self._pw.chromium.launch(
            headless=True,
            args=self._BROWSER_ARGS,
        )
        self._new_context()
        self._seed_cookies()

    def stop(self) -> None:
        if self._browser:
            try:
                self._browser.close()
            except Exception:
                pass
        if self._pw:
            try:
                self._pw.stop()
            except Exception:
                pass
        log.info("Browser stopped.")

    # ── Cookie management ──────────────────────────────────────────────────────

    def _new_context(self) -> None:
        if self._context:
            try:
                self._context.close()
            except Exception:
                pass
        self._context = self._browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1920, "height": 1080},
            locale="en-IN",
            timezone_id="Asia/Kolkata",
            extra_http_headers={
                "Accept-Language": "en-IN,en-US;q=0.9,en;q=0.8",
                "sec-ch-ua":          '"Chromium";v="124", "Google Chrome";v="124"',
                "sec-ch-ua-mobile":   "?0",
                "sec-ch-ua-platform": '"Windows"',
            },
        )
        # Remove webdriver fingerprint
        self._context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            Object.defineProperty(navigator, 'plugins',   {get: () => [1,2,3,4,5]});
            Object.defineProperty(navigator, 'languages', {get: () => ['en-IN','en']});
            window.chrome = {runtime: {}};
        """)

    def _seed_cookies(self) -> None:
        """
        Three-page visit sequence to acquire all NSE session cookies:
          1. Homepage  →  bm_sz, ak_bmsc (Akamai)
          2. All-Reports page  →  NSE navigation context
          3. Deals page  →  nsit, nseappid cookies needed for the API
        """
        page = self._context.new_page()
        try:
            for url, wait_ms in [
                (NSE_HOME,        2500),
                (NSE_ALL_REPORTS, 2000),
                (NSE_DEALS_PAGE,  3000),
            ]:
                try:
                    page.goto(url, wait_until="domcontentloaded", timeout=30_000)
                    page.wait_for_timeout(wait_ms)
                    log.debug("Seeded %s", url)
                except PWTimeout:
                    log.warning("Timeout seeding %s — continuing", url)
        finally:
            page.close()

        self._export_cookies()
        self._cookie_ts = time.time()
        log.info(
            "Cookies refreshed — %s",
            [c["name"] for c in self._context.cookies()],
        )

    def _export_cookies(self) -> None:
        """Copy Playwright cookies into a requests.Session."""
        self._api_session = requests.Session()
        self._api_session.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept":          "application/json, text/plain, */*",
            "Accept-Language": "en-IN,en-US;q=0.9,en;q=0.8",
            "Referer":         NSE_DEALS_PAGE,
            "sec-fetch-dest":  "empty",
            "sec-fetch-mode":  "cors",
            "sec-fetch-site":  "same-origin",
        })
        for c in self._context.cookies():
            self._api_session.cookies.set(
                c["name"], c["value"],
                domain=c.get("domain", "").lstrip("."),
            )

    def ensure_fresh(self) -> None:
        """Refresh cookies if TTL has elapsed."""
        if time.time() - self._cookie_ts > COOKIE_TTL:
            log.info("Cookie TTL elapsed — refreshing …")
            self._new_context()
            self._seed_cookies()

    # ── API fetch ──────────────────────────────────────────────────────────────

    def fetch_deals(self, option_type: str, nse_date: str) -> pd.DataFrame | None:
        """
        Download the full CSV for a given deal type and date.

        Uses the &csv=true endpoint which returns ALL records with no server-side
        cap (the JSON API is hard-capped at 70 records regardless of parameters).

        nse_date format: DD-MM-YYYY
        Returns None when the exchange has not yet published data for this date.
        """
        url = (
            f"{NSE_DEALS_API}"
            f"?optionType={option_type}&from={nse_date}&to={nse_date}&csv=true"
        )
        try:
            resp = self._api_session.get(url, timeout=20)

            if resp.status_code == 401:
                log.warning("401 on %s — refreshing cookies", option_type)
                self._seed_cookies()
                resp = self._api_session.get(url, timeout=20)

            resp.raise_for_status()

            if not resp.content:
                log.debug("%s: empty response — not yet published for %s", option_type, nse_date)
                return None

            # CSV is UTF-8 with BOM; requests decompresses gzip automatically
            df = pd.read_csv(
                io.BytesIO(resp.content),
                encoding="utf-8-sig",   # strips BOM
                skipinitialspace=True,
            )

            # Strip whitespace from column names (NSE adds trailing spaces)
            df.columns = df.columns.str.strip()

            if df.empty:
                log.debug("%s: empty CSV for %s", option_type, nse_date)
                return None

            log.info("%s: %d records for %s (full CSV)", option_type, len(df), nse_date)
            return df

        except Exception as exc:
            log.warning("fetch_deals(%s, %s) failed: %s", option_type, nse_date, exc)
            return None


# ══════════════════════════════════════════════════════════════════════════════
# Data processing
# ══════════════════════════════════════════════════════════════════════════════

def normalise(df: pd.DataFrame) -> pd.DataFrame:
    """
    Map CSV column names → display names, keep required fields, compute Traded Value.
    Handles Indian-format numbers (e.g. "12,80,967" → 1280967).
    """
    df = df.rename(columns=CSV_COL_MAP)
    keep = [c for c in REQUIRED_COLS if c in df.columns]
    df   = df[keep].copy()

    for col in ("Quantity Traded", "Traded Price"):
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace(",", "", regex=False),
                errors="coerce",
            )

    if {"Quantity Traded", "Traded Price"}.issubset(df.columns):
        df["Traded Value"] = df["Quantity Traded"] * df["Traded Price"]

    return df


def content_hash(df: pd.DataFrame) -> str:
    if df.empty:
        return ""
    return hashlib.md5(
        pd.util.hash_pandas_object(df, index=False).values.tobytes()
    ).hexdigest()


# ══════════════════════════════════════════════════════════════════════════════
# Excel builder
# ══════════════════════════════════════════════════════════════════════════════

def _write_sheet(ws, df: pd.DataFrame, label: str) -> None:
    if df.empty:
        cell = ws.cell(1, 1, f"No {label} data published yet for this session.")
        cell.font = Font(italic=True, color="888888", name="Calibri", size=11)
        return

    cols = [c for c in OUTPUT_COLS if c in df.columns]
    df   = df[cols]

    # Header
    ws.append(cols)
    for cell in ws[1]:
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = CELL_BORDER
    ws.row_dimensions[1].height = 28

    # Data rows
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        ws.append(row)
        row_fill = ALT_FILL if r_idx % 2 == 0 else None
        for c_idx, cell in enumerate(ws[r_idx], start=1):
            cell.font      = BODY_FONT
            cell.border    = CELL_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if row_fill:
                cell.fill = row_fill
            col_name = cols[c_idx - 1]
            if col_name == "Traded Value":
                cell.number_format = INR_FMT
            elif col_name in ("Quantity Traded", "Traded Price"):
                cell.number_format = "#,##0.00"

    # Auto column widths
    for c_idx, col_name in enumerate(cols, start=1):
        letter   = get_column_letter(c_idx)
        col_vals = [str(cell.value) for cell in ws[letter] if cell.value is not None]
        max_len  = max((len(v) for v in col_vals), default=10)
        ws.column_dimensions[letter].width = min(max_len + 4, 52)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_excel(bulk_df: pd.DataFrame, block_df: pd.DataFrame, file_date: str) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb.create_sheet("Bulk Deals"),  bulk_df,  "Bulk Deal")
    _write_sheet(wb.create_sheet("Block Deals"), block_df, "Block Deal")
    path = OUTPUT_DIR / f"NSEDEALS_{file_date}.xlsx"
    wb.save(path)
    log.info("Excel saved → %s  (bulk=%d, block=%d)", path, len(bulk_df), len(block_df))
    return path


# ══════════════════════════════════════════════════════════════════════════════
# Email
# ══════════════════════════════════════════════════════════════════════════════

def send_email(path: Path) -> None:
    if not GMAIL_USER or not GMAIL_PASSWORD:
        log.warning("Gmail credentials not set — email skipped.")
        return

    msg             = EmailMessage()
    msg["Subject"]  = f"NSE Deals — {path.stem}"
    msg["From"]     = GMAIL_USER
    msg["To"]       = ", ".join(EMAIL_TO)
    msg.set_content("Please find attached files")
    msg.add_alternative("<html><body><p>Please find attached files</p></body></html>", subtype="html")

    with open(path, "rb") as fh:
        msg.add_attachment(
            fh.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=path.name,
        )

    try:
        ctx = ssl.create_default_context()
        with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
            smtp.ehlo()
            smtp.starttls(context=ctx)
            smtp.login(GMAIL_USER, GMAIL_PASSWORD)
            smtp.send_message(msg)
        log.info("Email sent → %s", EMAIL_TO)
    except Exception as exc:
        log.error("Email failed: %s", exc)


# ══════════════════════════════════════════════════════════════════════════════
# Monitor loop
# ══════════════════════════════════════════════════════════════════════════════

_shutdown = False

def _handle_signal(sig, frame):
    global _shutdown
    log.info("Shutdown signal received — exiting after current cycle.")
    _shutdown = True

signal.signal(signal.SIGINT,  _handle_signal)
signal.signal(signal.SIGTERM, _handle_signal)


def run_monitor() -> None:
    session = NSEBrowserSession()
    session.start()

    seen: dict[str, str] = {"bulk": "", "block": ""}
    log.info("Monitor running — poll every %ds, cookie refresh every %ds", POLL_INTERVAL, COOKIE_TTL)

    try:
        while not _shutdown:
            session.ensure_fresh()

            nse_date  = datetime.datetime.now(IST).strftime("%d-%m-%Y")
            file_date = datetime.datetime.now(IST).strftime("%d%m%y")

            raw_bulk  = session.fetch_deals("bulk_deals",  nse_date)
            raw_block = session.fetch_deals("block_deals", nse_date)

            bulk_df  = normalise(raw_bulk)  if raw_bulk  is not None else pd.DataFrame()
            block_df = normalise(raw_block) if raw_block is not None else pd.DataFrame()

            h_bulk  = content_hash(bulk_df)
            h_block = content_hash(block_df)

            new_bulk  = h_bulk  and h_bulk  != seen["bulk"]
            new_block = h_block and h_block != seen["block"]

            if new_bulk or new_block:
                log.info("New data — bulk_new=%s  block_new=%s", new_bulk, new_block)
                path = build_excel(bulk_df, block_df, file_date)
                send_email(path)
                seen["bulk"]  = h_bulk
                seen["block"] = h_block
            else:
                nxt = (datetime.datetime.now() + datetime.timedelta(seconds=POLL_INTERVAL)).strftime("%H:%M:%S")
                log.info("No new data for %s (IST).  Next check at %s.", nse_date, nxt)

            for _ in range(POLL_INTERVAL):
                if _shutdown:
                    break
                time.sleep(1)

    finally:
        session.stop()


def run_once(target_date: str | None = None) -> None:
    """
    Single fetch-and-email run. target_date = 'DD-MM-YYYY' (IST).
    Defaults to today (IST). Use for testing or manual back-fills.
    """
    nse_date  = target_date or datetime.datetime.now(IST).strftime("%d-%m-%Y")
    file_date = datetime.datetime.strptime(nse_date, "%d-%m-%Y").strftime("%d%m%y")

    session = NSEBrowserSession()
    session.start()
    try:
        raw_bulk  = session.fetch_deals("bulk_deals",  nse_date)
        raw_block = session.fetch_deals("block_deals", nse_date)

        bulk_df  = normalise(raw_bulk)  if raw_bulk  is not None else pd.DataFrame()
        block_df = normalise(raw_block) if raw_block is not None else pd.DataFrame()

        log.info("Fetched — bulk=%d  block=%d", len(bulk_df), len(block_df))

        path = build_excel(bulk_df, block_df, file_date)
        send_email(path)
        return path, bulk_df, block_df
    finally:
        session.stop()


# ── Entry point ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if "--once" in sys.argv:
        # Optional date argument: python nse_deals_monitor.py --once 19-05-2026
        date_arg = next((a for a in sys.argv[1:] if a != "--once"), None)
        run_once(date_arg)
    else:
        run_monitor()
